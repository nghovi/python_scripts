#!/usr/bin/python
# -*- coding: utf-8 -*- 
"""This scripts contains function that help dealing with language file for android

    Created by vietnh
"""

import re
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

FILE_EN = '/Users/hviet/projects/atwork-adr/welfare-android-core/src/main/res/values/strings.xml'
FILE_JP = '/Users/hviet/projects/atwork-adr/welfare-android-core/src/main/res/values-ja/strings.xml'
FILE_VI = '/Users/hviet/projects/atwork-adr/welfare-android-core/src/main/res/values-vi/strings.xml'
FILE_LANGUAGE = '../data/language.xlsx'

# from android string resource files, export to excel file
# with key from en
def init_excel_file(file_en, file_jp, file_vi, file_out):
    wb = load_workbook(file_out)
    ws = wb.active

    fen = open(file_en, 'r')
    fjp = open(file_jp, 'r')
    fvi = open(file_vi, 'r')

    lines_jp = fjp.readlines()
    lines_vi = fvi.readlines()

    ws.cell(row = 1, column = 1).value = "Key"
    ws.cell(row = 1, column = 2).value = "English"
    ws.cell(row = 1, column = 3).value = "Japanese"
    ws.cell(row = 1, column = 4).value = "Vietnamese"


    row = 2
    col = 1
    for line_i in fen:
        # <string name="app_name">カレンダー</string>
        # <resources>
        # <!--Notifications-->
        if  re.search('<string', line_i): #if it's not comment line
            # print "line_i == " + line_i
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            # print "key == " + key
            value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
            ws.cell(row = row, column = col).value = key
            ws.cell(row = row, column = col + 1).value = value
            for japanese_string in lines_jp:
                if  re.search('<string', japanese_string): #if it's not comment line
                    ja_key = japanese_string[japanese_string.index('name="') + 6 : japanese_string.index('">')]
                    if key == ja_key:
                        ja_value = japanese_string[japanese_string.index('">') + 2 : japanese_string.index('</string>')]
                        ws.cell(row = row, column = col + 2).value = ja_value

            for vi_string in lines_vi:
                if  re.search('<string', vi_string): #if it's not comment line
                    vi_key = vi_string[vi_string.index('name="') + 6 : vi_string.index('">')]
                    if key == vi_key:
                        vi_value = vi_string[vi_string.index('">') + 2 : vi_string.index('</string>')]
                        ws.cell(row = row, column = col + 3).value = vi_value

            row = row + 1

    fen.close()
    fjp.close()
    fvi.close()

    wb.save(FILE_LANGUAGE)
#end of init_excel_file

def append_to_row(rows, key, value, language):
    existed = False
    for row in rows:
        if row['key'] == key:
            row[language] = value
            existed = True
    if not existed:
        row = {'key': key}
        row[language] = value
        rows.append(row)

# get data from android language files, update existing excel file
def add_new_keys_to_excel_file(file_en, file_jp, file_vi, file_out):
    wb = load_workbook(file_out)
    ws = wb.active

    fen = open(file_en, 'r')
    fjp = open(file_jp, 'r')
    fvi = open(file_vi, 'r')

    wb = load_workbook(filename = FILE_LANGUAGE)
    ws = wb.active

    rows = []

    for line_i in fen:
        if  re.search('<string', line_i): #if it's not comment line
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key not in map(lambda x: x.value, ws['A']):
                value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
                append_to_row(rows, key, value, 'en')

    for line_i in fjp:
        if  re.search('<string', line_i): #if it's not comment line
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key not in map(lambda x: x.value, ws['A']):
                value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
                append_to_row(rows, key, value, 'jp')

    for line_i in fvi:
        if  re.search('<string', line_i): #if it's not comment line
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key not in map(lambda x: x.value, ws['A']):
                value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
                append_to_row(rows, key, value, 'vi')

    next_row_in_excel = len(ws['A']) + 1

    # print "next_row_in_excel" , next_row_in_excel
    # print rows

    font = Font(color=colors.RED)

    for row in rows:
        value_en, value_jp, value_vi = "", "", ""
        if 'en' in row:
            value_en = row['en']
        if 'jp' in row:
            value_jp = row['jp']
        if 'vi' in row:
            value_vi = row['vi']

        ws.cell(row = next_row_in_excel, column = 1).value = row['key']
        ws.cell(row = next_row_in_excel, column = 1).font = font
        ws.cell(row = next_row_in_excel, column = 2).value = value_en
        ws.cell(row = next_row_in_excel, column = 2).font = font
        ws.cell(row = next_row_in_excel, column = 3).value = value_jp
        ws.cell(row = next_row_in_excel, column = 3).font = font
        ws.cell(row = next_row_in_excel, column = 4).value = value_vi
        ws.cell(row = next_row_in_excel, column = 4).font = font

        next_row_in_excel += 1

    fen.close()
    fjp.close()
    fvi.close()

    wb.save(FILE_LANGUAGE)


def migrate(f_in, copy, target_key, value):
    if not value:
        value = ""
    existed = False
    for line_i in f_in:
        if  re.search('<string', line_i): #if it's not comment line
            print "line_i == " + line_i
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key == target_key:
                existed = True
                string_line = '    <string name="' + key + '">' + value + '</string>' + "\n"
                copy.append(string_line)

    if not existed:
        # print " key = " + target_key + ": value = " + value
        string_line = '    <string name="' + target_key + '">' + value + '</string>' + "\n"
        copy.append(string_line)
        #todo create key

def rewrite_to_file(f_out, data):
    f_out.seek(0)
    f_out.truncate()
    for li in data:
        f_out.write(li.encode('utf-8'))

# Import language from excel files to android language files
def migrate_languages():
    wb = load_workbook(filename = FILE_LANGUAGE, read_only=True)
    ws = wb.active

    fen = open(FILE_EN, 'r+')
    fjp = open(FILE_JP, 'r+')
    fvi = open(FILE_VI, 'r+')

    copy_en = ["<resources>\n"]
    copy_jp = ["<resources>\n"]
    copy_vi = ["<resources>\n"]

    for idx, row in enumerate(ws.rows):
        if idx >= 2: #ignore first row
            key = row[0].value
            en_value = row[1].value
            ja_value = row[2].value
            vi_value = row[3].value

            migrate(fen, copy_en, key, en_value)
            migrate(fjp, copy_jp, key, ja_value)
            migrate(fvi, copy_vi, key, vi_value)


    copy_en.append("</resources>")
    copy_jp.append("</resources>")
    copy_vi.append("</resources>")

    rewrite_to_file(fen, copy_en)
    rewrite_to_file(fjp, copy_jp)
    rewrite_to_file(fvi, copy_vi)

    fen.close()
    fjp.close()
    fvi.close()

def get_elv_id(elv_type, elv_no):
    if elv_type == 'main':
        if elv_no == 1:
            return 1;
        if elv_no == 2:
            return 2;
        if elv_no == 3:
            return 3;
        if elv_no == 4:
            return 4;
        if elv_no == 5:
            return 5;
    elif elv_type == 'people':
        if elv_no == 1:
            return 6;
        if elv_no == 2:
            return 7;
        if elv_no == 3:
            return 8;
        if elv_no == 4:
            return 9;
        if elv_no == 5:
            return 10;
    elif elv_type == 'temp':
        if elv_no == 1:
            return 11;
        if elv_no == 2:
            return 12;
        if elv_no == 3:
            return 13;
        if elv_no == 4:
            return 14;
        if elv_no == 5:
            return 15;


if __name__ == '__main__':
    init_excel_file(FILE_EN, FILE_JP, FILE_VI, FILE_LANGUAGE)
    add_new_keys_to_excel_file(FILE_EN, FILE_JP, FILE_VI, FILE_LANGUAGE)
    migrate_languages()
    



    # shift_id = 100000
    # pattern_id = 263;


    #delete from sw_working_time_checkin_history;
    #delete from sw_working_time_checkin;
    #delete from sw_working_time_daily;



    # for user in users:
    #     for idx, start_time in enumerate(start_times):
            
    #         end_time = end_times[idx]
    #         working_date = working_dates[idx]

    #         print "INSERT INTO `sw_work_shift_plan` (`table_id`, `company_id`, `user_id`, `pattern_id`, `work_date`, `location`, `gps_longitude`, `gps_latitude`, `gps_use`, `create_date`, `create_user`, `update_date`, `update_user`, `delete_flag`) VALUES (" + str(shift_id) + ", " + str(company_id) + ", " + str(user) + ", " + str(pattern_id) + ", '" + working_date + "', NULL, NULL, NULL, 0, '" + start_time + "', " + str(user) + ", '" + start_time + "', " + str(user) + ", 0);";


    #         #start date
    #         #print "INSERT INTO `sw_working_time_daily` (`daily_id`, `company_id`, `user_id`, `work_date`, `start_time`, `end_time`, `total_break`, `total_overtime`, `total_over_night`, `total_break_night`, `create_date`, `create_user`, `update_date`, `update_user`, `delete_flag`, `shift_id`, `gps_use`, `gps_longitude`, `gps_latitude`, `address`, `total_worktime`, `total_shorttime`) VALUES (" + str(daily_id) + ", " + str(company_id) + ", " + str(user) + ", '" + working_date + "', '" + start_time + "', '" + end_time + "', NULL, NULL, NULL, NULL, '" + start_time + "', " + str(user) + ", '" + start_time + "', " + str(user) + ", 0, NULL, NULL, NULL, NULL, NULL, NULL, NULL);";
    #         #print "INSERT INTO `sw_working_time_checkin` (`table_id`, `daily_id`, `checkin_time`, `from_device`, `location`, `note`, `work_date`, `work_type`, `create_date`, `create_user`, `update_date`, `update_user`, `delete_flag`, `type_id`, `approval_status`, `request_content`, `comment_approve`, `checkin_time_history`, `work_type_history`, `checkin_location_history`) VALUES (" + str(checkin_id) + ", " + str(daily_id) + ", '" + start_time + "', 'WEB', NULL, '', '" + working_date + "', 'ST', '" + start_time + "', " + str(user) + ", '" + start_time + "'," + str(user) + ", 0, NULL, NULL, NULL, NULL, NULL, NULL, NULL);";
    #         #print "INSERT INTO `sw_working_time_checkin_history` (`table_id`, `checkin_id`, `action_key`, `create_date`, `create_user`, `update_date`, `update_user`, `delete_flag`, `user_approve`, `comment_history`, `checkin_time_history`, `request_content_history`, `work_type`, `checkin_location_history`) VALUES (" + str(checkin_id_history) + ", " + str(checkin_id) + ", NULL, '" + start_time + "', " + str(user) + ", '" + start_time + "', " + str(user) + ", 0, + " + str(user) + ", NULL, '" + start_time + "', NULL, 'ST', NULL);";

    #         checkin_id += 1
    #         checkin_id_history += 1

    #         #end
    #         #print "INSERT INTO `sw_working_time_checkin` (`table_id`, `daily_id`, `checkin_time`, `from_device`, `location`, `note`, `work_date`, `work_type`, `create_date`, `create_user`, `update_date`, `update_user`, `delete_flag`, `type_id`, `approval_status`, `request_content`, `comment_approve`, `checkin_time_history`, `work_type_history`, `checkin_location_history`) VALUES (" + str(checkin_id) + ", " + str(daily_id) + ", '" + end_time + "', 'WEB', NULL, '', '" + working_date + "', 'ET', '" + end_time + "', " + str(user) + ", '" + end_time + "'," + str(user) + ", 0, NULL, NULL, NULL, NULL, NULL, NULL, NULL);";
    #         #print "INSERT INTO `sw_working_time_checkin_history` (`table_id`, `checkin_id`, `action_key`, `create_date`, `create_user`, `update_date`, `update_user`, `delete_flag`, `user_approve`, `comment_history`, `checkin_time_history`, `request_content_history`, `work_type`, `checkin_location_history`) VALUES (" + str(checkin_id_history) + ", " + str(checkin_id) + ", NULL, '" + end_time + "', " + str(user) + ", '" + end_time + "', " + str(user) + ", 0, + " + str(user) + ", NULL, '" + end_time + "', NULL, 'ET', NULL);";

    #         daily_id += 1
    #         checkin_id += 1
    #         checkin_id_history += 1
    #         shift_id += 1

#################################################################


# from openpyxl import Workbook
# wb = Workbook()

# # grab the active worksheet
# ws = wb.active

# # Data can be assigned directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("sample.xlsx")